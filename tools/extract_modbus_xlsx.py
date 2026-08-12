import openpyxl, json, re, sys

def norm(s):
    return re.sub(r'\s+',' ',str(s)).strip().lower() if s is not None else ''

def extrae(ws):
    """Localiza la fila de cabecera por la celda 'Variable name' y mapea columnas por su rótulo.
       Las hojas del R7 tienen distinto número de columnas de relleno a la izquierda, así que
       fijar índices a mano se rompe en cuanto cambia una."""
    hdr=None
    for r in range(1, min(ws.max_row,15)+1):
        for c in range(1, min(ws.max_column,20)+1):
            if norm(ws.cell(r,c).value)=='variable name':
                hdr=r; break
        if hdr: break
    if not hdr: return None,None
    cols={}
    for c in range(1, min(ws.max_column,20)+1):
        k=norm(ws.cell(hdr,c).value)
        if k: cols[k]=c
    filas=[]
    for r in range(hdr+1, ws.max_row+1):
        g=lambda k: ws.cell(r,cols[k]).value if k in cols else None
        nombre=g('variable name')
        if nombre is None or str(nombre).strip()=='': continue
        filas.append({
            'addr':   g('register address'),
            'offset': g('offset'),
            'acc':    (str(g('register access')).strip() if g('register access') is not None else ''),
            'bits':   (str(g('(msb..lsb)')).strip() if g('(msb..lsb)') is not None else ''),
            'tipo':   (str(g('type')).strip() if g('type') is not None else ''),
            'nombre': str(nombre).strip(),
            'desc':   re.sub(r'\s+',' ',str(g('variable description'))).strip() if g('variable description') is not None else '',
            'rango':  str(g('range')).strip() if g('range') is not None else '',
            'unidad': str(g('unit')).strip() if g('unit') is not None else '',
            'defecto':str(g('default value')).strip() if g('default value') is not None else '',
        })
    return cols, filas

def extrae_overview(ws):
    """Hoja «Overview» del R7: el reparto COMPLETO del espacio de direcciones, con los bloques
       reservados y los huecos libres. Es lo que permite decir «esta direccion es un hueco
       reservado del bloque» en vez de un «no existe» a secas."""
    bloques=[]
    for r in range(2, ws.max_row+1):
        nom=ws.cell(r,2).value
        a1,a2=ws.cell(r,3).value, ws.cell(r,4).value
        if nom is None or not isinstance(a1,(int,float)) or not isinstance(a2,(int,float)): continue
        if a2 < a1: continue
        b={'nombre':re.sub(r'\s+',' ',str(nom)).strip(),'de':int(a1),'a':int(a2),
           'tam':ws.cell(r,5).value,'unidades':ws.cell(r,6).value}
        l1,l2=ws.cell(r,10).value, ws.cell(r,11).value
        if isinstance(l1,(int,float)) and isinstance(l2,(int,float)) and l2>=l1:
            b['libre']=[int(l1),int(l2)]
        bloques.append(b)
    return bloques

out={}
for f,lab in [('/root/.claude/uploads/73817923-79b4-5d11-9e5e-27a79f17b20a/32737926-NCU_Modbus_Map_R7_1.xlsx','ncu_r7'),
              ('/root/.claude/uploads/73817923-79b4-5d11-9e5e-27a79f17b20a/91ba946e-250506_HSU_Modbus_Map_R23.xlsx','hsu_r23')]:
    wb=openpyxl.load_workbook(f,data_only=True)
    out[lab]={}
    for ws in wb.worksheets:
        cols,filas=extrae(ws)
        if filas is None:
            print(f'  {lab}/{ws.title}: SIN cabecera "Variable name" (no es hoja de registros)')
            continue
        out[lab][ws.title]=filas
        print(f'  {lab}/{ws.title}: {len(filas)} filas · columnas {sorted(cols.keys())}')
# el reparto del espacio de direcciones va aparte: no es una tabla de registros
wb=openpyxl.load_workbook('/root/.claude/uploads/73817923-79b4-5d11-9e5e-27a79f17b20a/32737926-NCU_Modbus_Map_R7_1.xlsx',data_only=True)
out['bloques_r7']=extrae_overview(wb['Overview'])
print(f"  bloques del espacio de direcciones (hoja Overview): {len(out['bloques_r7'])}")

json.dump(out,open('/tmp/modbus_docs.json','w'),ensure_ascii=False,indent=1)
print('\nescrito /tmp/modbus_docs.json')
